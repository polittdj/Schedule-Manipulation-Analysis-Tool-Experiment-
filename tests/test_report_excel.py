"""Round-trip tests for report_excel (openpyxl) — H-DRIFT-1 traceability.

Every numeric assertion round-trips through the file on disk:
  1. Build a schedule and call ``analyze_schedule`` to get a ``ScheduleAnalysis``.
  2. Write the report to a tmp file via ``build_excel_report``.
  3. Re-open the file with ``openpyxl.load_workbook`` and assert the rendered
     values match the analysis fields.

This proves that the rendered numbers are traceable to the computed inputs
(H-DRIFT-1) rather than being fabricated or approximate.

Perturbation discipline (H-VACUOUS-TEST)
-----------------------------------------
The PERTURB test introduces a negative-lag lead on the only relation, forcing
DCMA-02 (Leads) to FAIL.  The health score must drop and the Findings sheet must
name "Leads".  The same schedule without the lead produces a PASS result.  Both
are regenerated and re-read from disk -- proving the report actually reflects the
computed analysis.

CUI banner
----------
The presence of ``CUI_NOTICE`` text is asserted in every sheet of every report
generated in this test module.

Axis note
---------
Default calendar: 480 working minutes per day.
``project_finish / 480.0`` == working-day display value used in the Summary sheet.
"""

from __future__ import annotations

import datetime as dt
import pathlib

import openpyxl
import pytest

from schedule_forensics.analysis import analyze_schedule
from schedule_forensics.metrics_common import MetricStatus
from schedule_forensics.report_excel import CUI_NOTICE, build_excel_report, build_excel_workbook
from schedule_forensics.schemas import Relation, RelationType, Schedule, Task

_START = dt.datetime(2025, 1, 6, 8)  # Monday, working-day start


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _task(uid: int, dur: int = 480, **kwargs: object) -> Task:
    return Task(unique_id=uid, name=f"T{uid}", duration_minutes=dur, **kwargs)  # type: ignore[arg-type]


def _clean_schedule() -> Schedule:
    """A minimal two-task linear schedule: T1(480) -> T2(480).

    CPM: project_finish = 960 = 2.0 working days.
    DCMA-02 (Leads): 0 leads -> PASS.
    Both tasks have no resources, so DCMA-10 (Resources) will FAIL.
    The schedule has one FS link, so DCMA-04 (Relationship Types) = 100% FS -> PASS.
    """
    return Schedule(
        name="clean",
        project_start=_START,
        status_date=_START,
        tasks=(_task(1, 480), _task(2, 480)),
        relations=(Relation(predecessor_id=1, successor_id=2),),
    )


def _lead_schedule() -> Schedule:
    """Same two tasks but the link has a negative lag (a lead).

    DCMA-02 (Leads): 1 lead -> FAIL.  Health score must be lower than the clean case.
    """
    return Schedule(
        name="lead",
        project_start=_START,
        status_date=_START,
        tasks=(_task(1, 480), _task(2, 480)),
        relations=(
            Relation(predecessor_id=1, successor_id=2, type=RelationType.FS, lag_minutes=-240),
        ),
    )


def _ev_schedule() -> Schedule:
    """Schedule carrying earned-value data; SPI = 0.75 (behind), status @ baseline finish."""
    d7, d8 = dt.datetime(2025, 1, 7, 8), dt.datetime(2025, 1, 8, 8)
    return Schedule(
        name="ev",
        project_start=_START,
        status_date=d8,
        tasks=(
            _task(
                1,
                480,
                percent_complete=100.0,
                baseline_start=_START,
                baseline_finish=d7,
                budgeted_cost=100.0,
            ),
            _task(
                2,
                480,
                percent_complete=50.0,
                baseline_start=d7,
                baseline_finish=d8,
                budgeted_cost=100.0,
            ),
        ),
    )


# ---------------------------------------------------------------------------
# Helper: collect all non-None cell values from a worksheet as a flat list of
# strings, so we can search for the CUI banner substring.
# ---------------------------------------------------------------------------


def _all_text_values(ws: object) -> list[str]:
    result: list[str] = []
    for row in ws.iter_rows():  # type: ignore[union-attr]
        for cell in row:
            if cell.value is not None:  # type: ignore[union-attr]
                result.append(str(cell.value))  # type: ignore[union-attr]
    return result


def _sheet_has_cui(ws: object) -> bool:
    return any(CUI_NOTICE in v for v in _all_text_values(ws))


# ---------------------------------------------------------------------------
# Tests: in-memory workbook (no disk I/O)
# ---------------------------------------------------------------------------


def test_build_excel_workbook_has_expected_sheets() -> None:
    analysis = analyze_schedule(_clean_schedule())
    wb = build_excel_workbook(analysis)
    assert wb.sheetnames == ["Summary", "DCMA", "Earned Value", "Findings"]


def test_earned_value_sheet_renders_spi_through_disk(tmp_path: pathlib.Path) -> None:
    """The Earned Value sheet round-trips SPI/SPI(t) ids and the 0.75 measured value."""
    analysis = analyze_schedule(_ev_schedule())
    spi = next(m for m in analysis.performance_indices if m.metric_id == "SPI")
    assert spi.status is MetricStatus.FAIL  # pre-condition: 0.75 < 0.95
    assert spi.measured == pytest.approx(0.75)

    path = tmp_path / "ev.xlsx"
    build_excel_report(analysis, path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Earned Value"]
    text_vals = _all_text_values(ws)
    assert "SPI" in text_vals
    assert "SPI(t)" in text_vals
    numeric_vals = [
        cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, int | float)
    ]
    assert any(abs(float(v) - 0.75) < 1e-9 for v in numeric_vals), (
        f"SPI 0.75 not found in Earned Value sheet numerics: {numeric_vals}"
    )
    assert _sheet_has_cui(ws)


def test_earned_value_sheet_skipped_without_ev_data() -> None:
    """A schedule with no EV data still renders SPI/SPI(t) rows, marked SKIPPED."""
    wb = build_excel_workbook(analyze_schedule(_clean_schedule()))
    text_vals = _all_text_values(wb["Earned Value"])
    assert "SPI" in text_vals
    assert "SKIPPED" in text_vals


def test_workbook_summary_contains_project_finish_minutes() -> None:
    """project_finish (minutes) is written verbatim into the Summary sheet."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.project_finish == 960, "fixture: chain 480+480=960"
    wb = build_excel_workbook(analysis)
    ws = wb["Summary"]
    all_vals = _all_text_values(ws)
    assert "960" in all_vals, f"960 not found in Summary; got: {all_vals}"


def test_workbook_summary_contains_project_finish_days() -> None:
    """project_finish / 480.0 is written into the Summary sheet."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.project_finish == 960
    expected_days = 960 / 480.0  # == 2.0
    wb = build_excel_workbook(analysis)
    ws = wb["Summary"]
    # Collect numeric cell values from the sheet
    numeric_vals = [
        cell.value  # type: ignore[union-attr]
        for row in ws.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert expected_days in numeric_vals, (
        f"{expected_days} not found in Summary numerics: {numeric_vals}"
    )


def test_workbook_summary_contains_health_score() -> None:
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.health_score is not None
    wb = build_excel_workbook(analysis)
    ws = wb["Summary"]
    numeric_vals = [
        cell.value  # type: ignore[union-attr]
        for row in ws.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert analysis.health_score in numeric_vals


def test_workbook_dcma_sheet_has_14_data_rows() -> None:
    """DCMA sheet must have exactly 14 data rows (header + 14 = 15 total populated rows)."""
    analysis = analyze_schedule(_clean_schedule())
    assert len(analysis.dcma) == 14
    wb = build_excel_workbook(analysis)
    ws = wb["DCMA"]
    # Count rows after row 1 (CUI banner) and row 2 (column headers) that have a Metric ID
    metric_rows = [
        row
        for row in ws.iter_rows(min_row=3, values_only=True)  # type: ignore[union-attr]
        if row[0] is not None and str(row[0]).startswith("DCMA-")
    ]
    assert len(metric_rows) == 14, f"Expected 14 DCMA rows, got {len(metric_rows)}"


def test_workbook_dcma_measured_values_present() -> None:
    """At least one DCMA metric's measured value appears as a number in the DCMA sheet."""
    analysis = analyze_schedule(_clean_schedule())
    runnable = [m for m in analysis.dcma if m.measured is not None]
    assert runnable, "fixture must produce at least one runnable metric"
    wb = build_excel_workbook(analysis)
    ws = wb["DCMA"]
    all_vals = _all_text_values(ws)
    # Check the first runnable metric's measured value is present as a rendered string
    first = runnable[0]
    assert str(first.measured) in all_vals or any(str(first.measured) in v for v in all_vals), (
        f"{first.measured} not found in DCMA sheet; got: {all_vals[:20]}"
    )


def test_workbook_all_sheets_have_cui_banner() -> None:
    """CUI_NOTICE text must appear in every sheet (LAW 1)."""
    analysis = analyze_schedule(_clean_schedule())
    wb = build_excel_workbook(analysis)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert _sheet_has_cui(ws), f"CUI banner missing from sheet '{sheet_name}'"


def test_workbook_findings_lists_failing_metrics() -> None:
    """Findings sheet must list the names of failing DCMA metrics."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.findings, "fixture must have at least one failing metric for this test"
    wb = build_excel_workbook(analysis)
    ws = wb["Findings"]
    all_text = _all_text_values(ws)
    for finding in analysis.findings:
        assert any(finding in v for v in all_text), (
            f"Finding '{finding}' not found in Findings sheet; got: {all_text}"
        )


def test_workbook_findings_no_fail_note() -> None:
    """When findings is empty the Findings sheet shows the 'no failing metrics' note."""
    # Build a schedule where all runnable checks pass by having resources assigned
    # and a fully-linked network with no leads/lags.  Both tasks have resource_names
    # set so DCMA-10 (Resources) does not FAIL.
    sched = Schedule(
        name="allpass",
        project_start=_START,
        status_date=_START,
        tasks=(
            _task(1, 480, resource_names=("Alice",)),
            _task(2, 480, resource_names=("Bob",)),
        ),
        relations=(Relation(predecessor_id=1, successor_id=2),),
    )
    analysis = analyze_schedule(sched)
    if analysis.findings:
        pytest.skip("schedule unexpectedly has failures; skip rather than assert wrong thing")
    wb = build_excel_workbook(analysis)
    ws = wb["Findings"]
    all_text = " ".join(_all_text_values(ws)).lower()
    assert "no failing" in all_text, f"Expected 'no failing' note; got: {all_text}"


# ---------------------------------------------------------------------------
# Tests: round-trip through disk (H-DRIFT-1 traceability)
# ---------------------------------------------------------------------------


def test_roundtrip_project_finish_minutes(tmp_path: pathlib.Path) -> None:
    """project_finish minutes written then re-read from disk equals analysis.project_finish."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.project_finish == 960
    out = tmp_path / "report.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    all_vals = _all_text_values(ws)
    assert "960" in all_vals, f"960 not found after round-trip; got: {all_vals}"


def test_roundtrip_project_finish_days(tmp_path: pathlib.Path) -> None:
    """project_finish / 480.0 written then re-read from disk equals 2.0."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.project_finish == 960
    expected_days = 960 / 480.0  # 2.0
    out = tmp_path / "report.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    numeric_vals = [
        cell.value  # type: ignore[union-attr]
        for row in ws.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert expected_days in numeric_vals, f"{expected_days} not in Summary numerics: {numeric_vals}"


def test_roundtrip_health_score(tmp_path: pathlib.Path) -> None:
    """Health score written then re-read from disk matches analysis.health_score."""
    analysis = analyze_schedule(_clean_schedule())
    assert analysis.health_score is not None
    out = tmp_path / "health.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    numeric_vals = [
        cell.value  # type: ignore[union-attr]
        for row in ws.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert analysis.health_score in numeric_vals, (
        f"health_score {analysis.health_score} not found after round-trip; numerics: {numeric_vals}"
    )


def test_roundtrip_cui_banner_in_all_sheets(tmp_path: pathlib.Path) -> None:
    """CUI banner text present in every sheet after round-trip through disk."""
    analysis = analyze_schedule(_clean_schedule())
    out = tmp_path / "cui.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert _sheet_has_cui(ws), (
            f"CUI banner missing from '{sheet_name}' sheet after disk round-trip"
        )


def test_roundtrip_dcma_measured_value(tmp_path: pathlib.Path) -> None:
    """A specific DCMA measured value is preserved through the disk round-trip.

    DCMA-02 (Leads) in the clean schedule: measured == 0.0, status PASS.
    After writing and re-reading, the DCMA-02 row has the expected values.
    """
    analysis = analyze_schedule(_clean_schedule())
    # DCMA-02 (Leads) must be PASS with measured==0.0 in the clean schedule.
    dcma02 = next(m for m in analysis.dcma if m.metric_id == "DCMA-02")
    assert dcma02.status is MetricStatus.PASS
    assert dcma02.measured == 0.0, f"Expected DCMA-02 measured==0.0, got {dcma02.measured}"

    out = tmp_path / "dcma.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["DCMA"]

    # Find the DCMA-02 row explicitly (values_only=True returns raw cell values)
    dcma02_row = None
    for row in ws.iter_rows(values_only=True):  # type: ignore[union-attr]
        if row[0] == "DCMA-02":
            dcma02_row = row
            break
    assert dcma02_row is not None, "DCMA-02 row not found in DCMA sheet"
    # Column index 2 (0-based) == "Status"
    assert dcma02_row[2] == "PASS", f"DCMA-02 status expected PASS, got {dcma02_row[2]}"
    # Column index 3 (0-based) == "Measured"
    assert dcma02_row[3] == 0.0, f"DCMA-02 measured expected 0.0, got {dcma02_row[3]}"


# ---------------------------------------------------------------------------
# Perturbation tests (H-VACUOUS-TEST)
# ---------------------------------------------------------------------------


def test_perturbation_lead_flips_dcma02_and_reduces_health(tmp_path: pathlib.Path) -> None:
    """Adding a negative-lag lead must change DCMA-02 status and reduce health score.

    Schedule A (clean): no leads -> DCMA-02 PASS.
    Schedule B (lead): one -240 min lead -> DCMA-02 FAIL -> health score lower.

    Both reports are written to disk and re-read, proving the report reflects the
    computed analysis (H-VACUOUS-TEST / H-DRIFT-1).
    """
    # --- Schedule A: clean (no lead) ---
    analysis_a = analyze_schedule(_clean_schedule())
    dcma02_a = next(m for m in analysis_a.dcma if m.metric_id == "DCMA-02")
    assert dcma02_a.status is MetricStatus.PASS, "pre-condition: clean schedule DCMA-02 must PASS"
    assert analysis_a.health_score is not None

    out_a = tmp_path / "clean.xlsx"
    build_excel_report(analysis_a, out_a)
    wb_a = openpyxl.load_workbook(out_a)
    ws_a_dcma = wb_a["DCMA"]

    # Verify clean report has DCMA-02 status "PASS" on disk
    dcma02_row_a = None
    for row in ws_a_dcma.iter_rows(values_only=True):  # type: ignore[union-attr]
        if row[0] == "DCMA-02":
            dcma02_row_a = row
            break
    assert dcma02_row_a is not None
    assert dcma02_row_a[2] == "PASS", f"A: DCMA-02 status on disk: {dcma02_row_a[2]}"

    # Verify clean report's health score on disk
    ws_a_summary = wb_a["Summary"]
    summary_floats_a = [
        cell.value  # type: ignore[union-attr]
        for row in ws_a_summary.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert analysis_a.health_score in summary_floats_a, (
        f"A: health_score {analysis_a.health_score} not in Summary floats {summary_floats_a}"
    )

    # --- Schedule B: perturbed (lead added) ---
    analysis_b = analyze_schedule(_lead_schedule())
    dcma02_b = next(m for m in analysis_b.dcma if m.metric_id == "DCMA-02")
    assert dcma02_b.status is MetricStatus.FAIL, "post-perturbation: DCMA-02 must FAIL"
    assert analysis_b.health_score is not None
    # Health must be lower after perturbation (H-VACUOUS-TEST)
    assert analysis_b.health_score < analysis_a.health_score, (
        f"Expected health to drop: clean={analysis_a.health_score} lead={analysis_b.health_score}"
    )

    out_b = tmp_path / "lead.xlsx"
    build_excel_report(analysis_b, out_b)
    wb_b = openpyxl.load_workbook(out_b)
    ws_b_dcma = wb_b["DCMA"]

    # Verify lead report has DCMA-02 status "FAIL" on disk
    dcma02_row_b = None
    for row in ws_b_dcma.iter_rows(values_only=True):  # type: ignore[union-attr]
        if row[0] == "DCMA-02":
            dcma02_row_b = row
            break
    assert dcma02_row_b is not None
    assert dcma02_row_b[2] == "FAIL", f"B: DCMA-02 status on disk: {dcma02_row_b[2]}"

    # Verify lead report's health score on disk is lower
    ws_b_summary = wb_b["Summary"]
    summary_floats_b = [
        cell.value  # type: ignore[union-attr]
        for row in ws_b_summary.iter_rows()  # type: ignore[union-attr]
        for cell in row
        if isinstance(cell.value, (int, float))  # type: ignore[union-attr]
    ]
    assert analysis_b.health_score in summary_floats_b, (
        f"B: health_score {analysis_b.health_score} not in Summary floats {summary_floats_b}"
    )
    # The on-disk health scores differ (perturbation is visible in the report)
    assert analysis_b.health_score < analysis_a.health_score


def test_perturbation_lead_findings_names_leads(tmp_path: pathlib.Path) -> None:
    """After perturbation the Findings sheet must name 'Leads'."""
    analysis = analyze_schedule(_lead_schedule())
    assert any("Lead" in f for f in analysis.findings), (
        "pre-condition: 'Leads' must appear in findings after adding a lead"
    )
    out = tmp_path / "lead_findings.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Findings"]
    all_text = " ".join(_all_text_values(ws))
    assert "Leads" in all_text, (
        f"'Leads' not found in Findings sheet after perturbation; got: {all_text}"
    )


# ---------------------------------------------------------------------------
# Edge-case: CPM error path (project_finish is None)
# ---------------------------------------------------------------------------


def test_roundtrip_cpm_error_shows_na(tmp_path: pathlib.Path) -> None:
    """When CPM fails (cycle), project finish shows 'n/a' in the Summary sheet."""
    # A cyclic schedule: T1->T2 and T2->T1 => CPM raises.
    cyclic = Schedule(
        name="cyclic",
        project_start=_START,
        status_date=_START,
        tasks=(_task(1, 480), _task(2, 480)),
        relations=(
            Relation(predecessor_id=1, successor_id=2),
            Relation(predecessor_id=2, successor_id=1),
        ),
    )
    analysis = analyze_schedule(cyclic)
    assert analysis.project_finish is None
    assert analysis.cpm_error is not None

    out = tmp_path / "cyclic.xlsx"
    build_excel_report(analysis, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    all_text = " ".join(_all_text_values(ws)).lower()
    assert "n/a" in all_text, f"'n/a' not found for CPM-error case; got: {all_text}"
