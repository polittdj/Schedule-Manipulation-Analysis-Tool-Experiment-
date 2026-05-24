"""Excel report generator for Schedule Forensics (openpyxl).

Public API
----------
``build_excel_workbook(analysis) -> Workbook``
    Build and return the workbook in memory (no disk I/O).

``build_excel_report(analysis, path) -> None``
    Build the workbook and write it to *path* as a ``.xlsx`` file.

Sheets
------
* **Summary** -- CUI banner; project finish (working-minute offset and working
  days = offset / 480.0); critical path ids; driving chain ids; health score;
  CPM error (if any).
* **DCMA** -- header row + 14 rows in id order. Columns: Metric ID, Name, Status,
  Measured, Threshold, Direction, Extension?, Source, Detail. Numeric measured /
  threshold are written as numbers; None -> blank.
* **Earned Value** -- header row + the SPI / SPI(t) indices (SKIPPED rows when the
  schedule carries no earned-value data; never fabricated).
* **Findings** -- one row per failing metric, or a "no failing metrics" note.
* **Risk (SRA)** -- present only when an :class:`~schedule_forensics.sra.SRAResult`
  is supplied: Monte-Carlo finish percentiles (P50/P80/P95, working days) +
  per-activity criticality index (reference-tool capability; the default duration
  spread that seeds it is the tool's own heuristic, captioned as such).
* **Version Diff** -- present only when one or more
  :class:`~schedule_forensics.diff_engine.VersionPairDiff` are supplied: objective
  consecutive-version deltas (added/deleted, became-critical/recovered, finish/float
  shifts in working days, logic add/remove). Measured facts, not a score.

CUI / LAW 1: no schedule data leaves the machine; all writes are local. The
``CUI_NOTICE`` constant is the single source of truth (in every sheet).

Units: working-minute offsets are divided by 480.0 for working days. This is a
display-only conversion; analysis values are never recomputed -- we render
exactly what ``ScheduleAnalysis`` holds (H-DRIFT-1 / LAW 2). Note: openpyxl
round-trips a whole float (e.g. ``2.0``) back as ``int 2``; that is the value
2 working days either way.
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from schedule_forensics.analysis import ScheduleAnalysis
from schedule_forensics.diff_engine import TaskDelta, VersionPairDiff
from schedule_forensics.metrics_common import MetricStatus
from schedule_forensics.sra import SRAResult
from schedule_forensics.trend_analysis import TrendReport

CUI_NOTICE: str = (
    "CONTROLLED UNCLASSIFIED INFORMATION (CUI) — generated locally; "
    "not for distribution. "
    "This report may contain CUI per NIST 800-171 / DFARS."
)

_MINUTES_PER_DAY: float = 480.0

_DCMA_HEADERS: list[str] = [
    "Metric ID",
    "Name",
    "Status",
    "Measured",
    "Threshold",
    "Direction",
    "Extension?",
    "Source",
    "Detail",
]

# Earned-value indices are never extensions, so no "Extension?" column.
_EV_HEADERS: list[str] = [
    "Metric ID",
    "Name",
    "Status",
    "Measured",
    "Threshold",
    "Direction",
    "Source",
    "Detail",
]


def _header_fill() -> PatternFill:
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _cui_fill() -> PatternFill:
    return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _status_fill(status: MetricStatus) -> PatternFill:
    if status is MetricStatus.PASS:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if status is MetricStatus.FAIL:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _add_cui_banner(ws: Worksheet, span: str) -> None:
    """Write the CUI banner across ``span`` (e.g. ``'A1:C1'``) on row 1."""
    cell = ws.cell(row=1, column=1, value=CUI_NOTICE)
    cell.font = Font(bold=True, color="FF0000")
    cell.fill = _cui_fill()
    ws.merge_cells(span)
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 32


def _autofit_columns(ws: Worksheet) -> None:
    """Expand each column to fit its widest content (approximate)."""
    for col_cells in ws.columns:
        first = col_cells[0]
        if first.column is None:
            continue
        col_letter = get_column_letter(first.column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)


def _build_summary_sheet(ws: Worksheet, analysis: ScheduleAnalysis) -> None:
    _add_cui_banner(ws, "A1:C1")

    for col, header in enumerate(("Field", "Value", "Notes"), start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()

    rows: list[tuple[str, object, str]] = []
    if analysis.project_finish is not None:
        rows.append(
            (
                "Project Finish (working minutes)",
                analysis.project_finish,
                "offset from project_start",
            )
        )
        rows.append(
            (
                "Project Finish (working days)",
                analysis.project_finish / _MINUTES_PER_DAY,
                "offset / 480.0",
            )
        )
    else:
        rows.append(("Project Finish (working minutes)", "n/a", "CPM could not be computed"))
        rows.append(("Project Finish (working days)", "n/a", "CPM could not be computed"))

    cp_str = ", ".join(str(uid) for uid in analysis.critical_path) or "(none)"
    rows.append(("Critical Path (UniqueIDs)", cp_str, "tasks with total_float <= 0, topo order"))
    dc_str = ", ".join(str(uid) for uid in analysis.driving_chain) or "(none)"
    rows.append(("Driving Chain (UniqueIDs)", dc_str, "binding-link back-trace to project finish"))

    if analysis.health_score is not None:
        rows.append(("Health Score", analysis.health_score, "% of runnable DCMA metrics that PASS"))
    else:
        rows.append(("Health Score", "n/a", "no runnable DCMA metrics"))

    if analysis.cpm_error:
        rows.append(("CPM Error", analysis.cpm_error, "CPM-dependent checks were SKIPPED"))

    for r_idx, (field, value, note) in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=field)
        ws.cell(row=r_idx, column=2, value=value)
        ws.cell(row=r_idx, column=3, value=note)

    _autofit_columns(ws)


def _build_dcma_sheet(ws: Worksheet, analysis: ScheduleAnalysis) -> None:
    _add_cui_banner(ws, "A1:I1")

    for col, header in enumerate(_DCMA_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()

    for r_idx, metric in enumerate(analysis.dcma, start=3):
        values: list[object] = [
            metric.metric_id,
            metric.name,
            str(metric.status),
            metric.measured,  # float or None -> blank when None
            metric.threshold,
            "" if metric.direction is None else str(metric.direction),
            "Yes" if metric.is_extension else "No",
            metric.source,
            metric.detail,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            if col == 3:
                cell.fill = _status_fill(metric.status)

    _autofit_columns(ws)


def _build_ev_sheet(ws: Worksheet, analysis: ScheduleAnalysis) -> None:
    """Earned-value indices (SPI / SPI(t)); SKIPPED rows when EV data is absent."""
    _add_cui_banner(ws, "A1:H1")

    for col, header in enumerate(_EV_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()

    if not analysis.performance_indices:
        ws.cell(row=3, column=1, value="no earned-value indices").font = Font(italic=True)
        _autofit_columns(ws)
        return

    for r_idx, metric in enumerate(analysis.performance_indices, start=3):
        values: list[object] = [
            metric.metric_id,
            metric.name,
            str(metric.status),
            metric.measured,  # float or None -> blank when None
            metric.threshold,
            "" if metric.direction is None else str(metric.direction),
            metric.source,
            metric.detail,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            if col == 3:
                cell.fill = _status_fill(metric.status)

    _autofit_columns(ws)


def _build_findings_sheet(ws: Worksheet, analysis: ScheduleAnalysis) -> None:
    _add_cui_banner(ws, "A1:B1")

    hdr = ws.cell(row=2, column=1, value="Failing Metric")
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = _header_fill()

    if analysis.findings:
        for r_idx, name in enumerate(analysis.findings, start=3):
            ws.cell(row=r_idx, column=1, value=name)
    else:
        ws.cell(row=3, column=1, value="no failing metrics").font = Font(italic=True)

    _autofit_columns(ws)


def _build_trends_sheet(ws: Worksheet, trends: TrendReport) -> None:
    """Multi-version trend analysis (TOOL-ORIGINAL EXTENSION): trajectory + float erosion."""
    _add_cui_banner(ws, "A1:F1")
    row = 2

    if trends.finish_days_net_change is not None:
        ws.cell(row=row, column=1, value="Finish drift (working days)").font = Font(bold=True)
        ws.cell(
            row=row,
            column=2,
            value=(
                f"{trends.finish_days_first:.1f} -> {trends.finish_days_last:.1f} "
                f"(net {trends.finish_days_net_change:+.1f})"
            ),
        )
        row += 2

    ws.cell(row=row, column=1, value="Version Trajectory").font = Font(bold=True)
    row += 1
    traj_headers = ["#", "Status date", "Finish (working days)", "Health %", "Band", "Critical"]
    for col, header in enumerate(traj_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
    row += 1
    for snap in trends.snapshots:
        ws.cell(row=row, column=1, value=snap.index + 1)
        ws.cell(
            row=row,
            column=2,
            value=snap.status_date.date().isoformat() if snap.status_date else "n/a",
        )
        ws.cell(
            row=row,
            column=3,
            value=snap.project_finish_days if snap.project_finish_days is not None else "n/a",
        )
        ws.cell(
            row=row, column=4, value=snap.health_score if snap.health_score is not None else "n/a"
        )
        ws.cell(row=row, column=5, value=snap.band)
        ws.cell(row=row, column=6, value=snap.n_critical)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Float Erosion bands (per task)").font = Font(bold=True)
    row += 1
    for band, count in trends.band_counts.items():
        ws.cell(row=row, column=1, value=band)
        ws.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    eroders = trends.worst_eroders(20)
    if eroders:
        ws.cell(row=row, column=1, value="Biggest float losers (worst first)").font = Font(
            bold=True
        )
        row += 1
        er_headers = [
            "UniqueID",
            "Earliest float (d)",
            "Latest float (d)",
            "Net change (d)",
            "Trend",
        ]
        for col, header in enumerate(er_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _header_fill()
        row += 1
        for trend in eroders:
            ws.cell(row=row, column=1, value=trend.unique_id)
            ws.cell(row=row, column=2, value=trend.earliest_float_days)
            ws.cell(row=row, column=3, value=trend.latest_float_days)
            ws.cell(row=row, column=4, value=trend.net_change_days)
            ws.cell(row=row, column=5, value=str(trend.trend))
            row += 1

    _autofit_columns(ws)


def _build_sra_sheet(ws: Worksheet, sra: SRAResult) -> None:
    """Monte-Carlo Schedule Risk Analysis: finish percentiles + criticality index.

    The Monte-Carlo SRA method (finish distribution + per-activity criticality
    index) is a reference-tool capability (Acumen Fuse Risk tab / Primavera Risk
    Analysis) and is NOT a tool-original extension. The DEFAULT duration spread
    used to seed it IS the tool's own heuristic (source-pending) -- captioned as
    such below, never presented as parity. Finish offsets are rendered in working
    days (offset / 480.0), consistent with the rest of the report."""
    _add_cui_banner(ws, "A1:C1")
    row = 2

    ws.cell(row=row, column=1, value="Monte-Carlo iterations").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sra.iterations)
    row += 1
    ws.cell(row=row, column=1, value="Deterministic finish (working days)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sra.deterministic_finish / _MINUTES_PER_DAY)
    row += 1
    ws.cell(row=row, column=1, value="Mean finish (working days)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sra.mean_finish / _MINUTES_PER_DAY)
    row += 2

    ws.cell(row=row, column=1, value="Finish percentiles (working days)").font = Font(bold=True)
    row += 1
    for col, header in enumerate(("Percentile", "Finish (working days)"), start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
    row += 1
    for label, offset in (("P50", sra.p50), ("P80", sra.p80), ("P95", sra.p95)):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=offset / _MINUTES_PER_DAY)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Criticality index (per activity)").font = Font(bold=True)
    row += 1
    ranked = sorted(sra.criticality_index.items(), key=lambda kv: (-kv[1], kv[0]))
    on_path = [(uid, ci) for uid, ci in ranked if ci > 0.0]
    if on_path:
        for col, header in enumerate(("UniqueID", "Criticality index (%)"), start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _header_fill()
        row += 1
        for uid, ci in on_path[:25]:
            ws.cell(row=row, column=1, value=uid)
            ws.cell(row=row, column=2, value=100.0 * ci)
            row += 1
    else:
        ws.cell(
            row=row, column=1, value="no activity reached the critical path in any iteration"
        ).font = Font(italic=True)
        row += 1
    row += 1

    note = ws.cell(
        row=row,
        column=1,
        value=(
            "SRA method (finish distribution + criticality index) mirrors Acumen Fuse / "
            "Primavera Risk Analysis (reference parity). The default duration spread that seeds "
            "it is the tool's own heuristic (source-pending); in an engagement the analyst "
            "supplies per-activity risk ranges. Distribution is over duration uncertainty only."
        ),
    )
    note.font = Font(italic=True)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    _autofit_columns(ws)


def _notable_deltas(pair: VersionPairDiff) -> list[TaskDelta]:
    """Task deltas where anything actually changed, most-moved (by finish shift) first."""
    changed = [
        d
        for d in pair.task_deltas
        if d.became_critical
        or d.recovered
        or d.finish_shift_minutes
        or d.total_float_delta_minutes
        or d.predecessors_added
        or d.predecessors_removed
    ]
    changed.sort(key=lambda d: (-abs(d.finish_shift_minutes), d.unique_id))
    return changed


def _build_diff_sheet(ws: Worksheet, diff: tuple[VersionPairDiff, ...]) -> None:
    """Objective version-to-version deltas (NOT an extension -- comparative facts).

    Per consecutive pair: the objective counts (added/deleted/became-critical/
    recovered/logic edits) and the most-moved task deltas. Shifts render in working
    days (offset / 480.0). These are measured facts (the Acumen ProjectTimeNow
    pattern), so no threshold is applied and nothing is flagged an extension."""
    _add_cui_banner(ws, "A1:F1")
    row = 3
    for pair in diff:
        header = (
            f"{pair.previous_status.date().isoformat()} -> {pair.current_status.date().isoformat()}"
        )
        ws.cell(row=row, column=1, value=header).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Tasks added / deleted")
        ws.cell(row=row, column=2, value=f"{len(pair.added_ids)} / {len(pair.deleted_ids)}")
        row += 1
        ws.cell(row=row, column=1, value="Became critical / recovered")
        n_bc = sum(1 for d in pair.task_deltas if d.became_critical)
        n_rec = sum(1 for d in pair.task_deltas if d.recovered)
        ws.cell(row=row, column=2, value=f"{n_bc} / {n_rec}")
        row += 1
        ws.cell(row=row, column=1, value="Logic links added + removed")
        n_logic = sum(
            len(d.predecessors_added) + len(d.predecessors_removed) for d in pair.task_deltas
        )
        ws.cell(row=row, column=2, value=n_logic)
        row += 1

        notable = _notable_deltas(pair)
        if notable:
            headers = [
                "UniqueID",
                "Finish shift (d)",
                "Float delta (d)",
                "Flag",
                "Preds +",
                "Preds -",
            ]
            for col, header_text in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header_text)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill()
            row += 1
            for d in notable:
                flag = (
                    "became critical" if d.became_critical else "recovered" if d.recovered else ""
                )
                ws.cell(row=row, column=1, value=d.unique_id)
                ws.cell(row=row, column=2, value=d.finish_shift_minutes / _MINUTES_PER_DAY)
                ws.cell(row=row, column=3, value=d.total_float_delta_minutes / _MINUTES_PER_DAY)
                ws.cell(row=row, column=4, value=flag)
                ws.cell(row=row, column=5, value=", ".join(str(p) for p in d.predecessors_added))
                ws.cell(row=row, column=6, value=", ".join(str(p) for p in d.predecessors_removed))
                row += 1
        else:
            ws.cell(row=row, column=1, value="no task-level changes").font = Font(italic=True)
            row += 1
        row += 1

    note = ws.cell(
        row=row,
        column=1,
        value=(
            "Objective version deltas (Acumen ProjectTimeNow / ProjectPreviousTimeNow comparative "
            "pattern): measured facts, not a score; no threshold is applied. Positive finish shift "
            "= task moved later; positive float delta = float gained."
        ),
    )
    note.font = Font(italic=True)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _autofit_columns(ws)


def build_excel_workbook(
    analysis: ScheduleAnalysis,
    trends: TrendReport | None = None,
    sra: SRAResult | None = None,
    diff: tuple[VersionPairDiff, ...] = (),
) -> Workbook:
    """Build and return an openpyxl ``Workbook``.

    Always writes Summary / DCMA / Earned Value / Findings sheets. When *trends* is
    given and spans 2+ versions, also appends a **Trends** sheet (tool-original
    extension: version trajectory + float-erosion bands). When *sra* is given,
    appends a **Risk (SRA)** sheet (Monte-Carlo finish percentiles + criticality
    index; reference-tool capability). When *diff* has any version pairs, appends a
    **Version Diff** sheet (objective consecutive-version deltas)."""
    wb = Workbook()
    summary = wb.active
    assert summary is not None  # a fresh Workbook always has an active sheet
    summary.title = "Summary"
    _build_summary_sheet(summary, analysis)
    _build_dcma_sheet(wb.create_sheet("DCMA"), analysis)
    _build_ev_sheet(wb.create_sheet("Earned Value"), analysis)
    _build_findings_sheet(wb.create_sheet("Findings"), analysis)
    if trends is not None and trends.n_versions > 1:
        _build_trends_sheet(wb.create_sheet("Trends"), trends)
    if sra is not None:
        _build_sra_sheet(wb.create_sheet("Risk (SRA)"), sra)
    if diff:
        _build_diff_sheet(wb.create_sheet("Version Diff"), diff)
    return wb


def build_excel_report(
    analysis: ScheduleAnalysis,
    path: str | os.PathLike[str],
    trends: TrendReport | None = None,
    sra: SRAResult | None = None,
    diff: tuple[VersionPairDiff, ...] = (),
) -> None:
    """Write an Excel ``.xlsx`` report for *analysis* to *path* (all data stays local)."""
    build_excel_workbook(analysis, trends=trends, sra=sra, diff=diff).save(path)
